import random

from openpyxl import Workbook

wb = Workbook(write_only=True)

print(wb.sheetnames)

# below comment code does not work in write_only mode
# ws = wb.active
#
# for row in range(1, 10):
#     for col in range(1, 10):
#         _ = ws.cell(row=row, column=col, value=random.randint(1, 1000))

ws1 = wb.create_sheet(title='new one')

# large dataset cost much cpu and memory
for irow in range(100):
    ws1.append([random.randint(1, 10000) for i in range(10)])

wb.save(filename='write.xlsx')
