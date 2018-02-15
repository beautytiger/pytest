from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

wb = load_workbook(filename='origin.xlsx', read_only=True)

sheets = wb.sheetnames
print(sheets)

ws = wb['文本数据']

print(ws.title)
print(ws.calculate_dimension())
print(ws.max_row)
print(ws.max_column)

print('convert column strings')
print(get_column_letter(4324))
print(column_index_from_string('AAA'))

print('get specific cell')
print(ws['A1'].value)
print(ws.cell(row=1, column=1).value)

exit()

# for row in ws.rows:
#     for cell in row:
#         val = cell.value
#         if isinstance(val, str):
#             val = val.replace('\n', '')
#         print(val, end=' ')
#     print('')

# this method is not recommended
# for i in range(1, 322):
#     for j in range(1, 18):
#         val = ws.cell(row=i, column=j).value
#         print(val, end=' ')
#     print('')
#     break

for row in ws.iter_rows(min_row=1, max_row=321, max_col=17):
    for index, cell in enumerate(row):
        val = cell.value
        if isinstance(val, str):
            val = val.replace('\n', '')
            val = val.replace('\t', '')
        print(index, val, end=' ')
    print('')

